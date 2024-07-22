import pandas as pd
from openpyxl import load_workbook
from tabulate import tabulate

# Fonction pour ajouter une entrée dans le stock
def add_entry(product_info):
    try:
        # Charger le workbook
        wb = load_workbook('Book.xlsx')

        # Accéder à la feuille 'Entrée' ou la créer si elle n'existe pas
        if 'Entrée' in wb.sheetnames:
            ws = wb['Entrée']
        else:
            ws = wb.create_sheet('Entrée')

            # Ajouter les en-têtes si la feuille est nouvellement créée
            headers = ['Num_entre', 'Reference_produit', 'quantite_entre']
            ws.append(headers)

        # Déterminer le prochain numéro d'entrée
        next_num_entre = ws.max_row + 1

        # Ajouter la nouvelle entrée dans le DataFrame
        new_entry = [next_num_entre, product_info['Produit'], product_info['Quantité']]
        ws.append(new_entry)

        # Sauvegarder le workbook
        wb.save('Book.xlsx')

        print(f"Entrée réussie pour le produit {product_info['Produit']} avec la quantité {product_info['Quantité']}.")
    except Exception as e:
        print(f"Erreur lors de l'ajout de l'entrée: {e}")

# Fonction pour ajouter une sortie dans le stock
def add_exit(product_info):
    try:
        # Charger le workbook
        wb = load_workbook('Book.xlsx')

        # Accéder à la feuille 'Sortie' ou la créer si elle n'existe pas
        if 'Sortie' in wb.sheetnames:
            ws = wb['Sortie']
        else:
            ws = wb.create_sheet('Sortie')

            # Ajouter les en-têtes si la feuille est nouvellement créée
            headers = ['Num_sortie', 'Reference_produit', 'quantite_sortie']
            ws.append(headers)

        # Déterminer le prochain numéro de sortie
        next_num_sortie = ws.max_row + 1

        # Ajouter la nouvelle sortie dans le DataFrame
        new_exit = [next_num_sortie, product_info['Produit'], product_info['Quantité']]
        ws.append(new_exit)

        # Sauvegarder le workbook
        wb.save('Book.xlsx')

        print(f"Sortie réussie pour le produit {product_info['Produit']} avec la quantité {product_info['Quantité']}.")
    except Exception as e:
        print(f"Erreur lors de l'ajout de la sortie: {e}")

# Fonction pour afficher le tableau de bord principal
def show_dashboard():
    while True:
        print("\nMenu:")
        print(tabulate([
            ['Produit'],
            ['Entrée'],
            ['Sortie'],
            ['Quitter']
        ], headers=['Option', 'Description'], tablefmt='fancy_grid'))

        choice = input("Choisissez une option: ")
        if choice.lower() == 'produit':
            # Ici vous pouvez ajouter la logique pour gérer les produits
            print("Gestion des produits...")
        elif choice.lower() == 'entrée':
            product_info = {}
            while True:
                product_info['Produit'] = input("Entrez le nom du produit (quittez avec 'q'): ")
                if product_info['Produit'].lower() == 'q':
                    break
                try:
                    product_info['Quantité'] = int(input("Entrez la quantité entrée: "))
                    if product_info['Quantité'] <= 0:
                        print("La quantité doit être supérieure à zéro.")
                        continue
                except ValueError:
                    print("Quantité invalide. Veuillez entrer un nombre entier positif.")
                    continue

                add_entry(product_info)
                break
        elif choice.lower() == 'sortie':
            product_info = {}
            while True:
                product_info['Produit'] = input("Entrez le nom du produit (quittez avec 'q'): ")
                if product_info['Produit'].lower() == 'q':
                    break
                try:
                    product_info['Quantité'] = int(input("Entrez la quantité sortie: "))
                    if product_info['Quantité'] <= 0:
                        print("La quantité doit être supérieure à zéro.")
                        continue
                except ValueError:
                    print("Quantité invalide. Veuillez entrer un nombre entier positif.")
                    continue

                add_exit(product_info)
                break
        elif choice.lower() == 'quitter':
            print("Au revoir!")
            break
        else:
            print("Option invalide. Veuillez choisir 'Produit', 'Entrée', 'Sortie' ou 'Quitter'.")

# Fonction principale pour démarrer le programme
def main():
    show_dashboard()

if __name__ == "__main__":
    main()
